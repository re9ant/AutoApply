import logging
import os
import shutil
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.config.settings import settings
from app.models.application import ApplicationRecord, ApplicationStatus
from app.models.excel_schema import DEFAULT_EXCEL_COLUMNS, get_column_alias_map

logger = logging.getLogger(__name__)


class ExcelTracker:
    """Synchronizes internal application states with an Excel (.xlsx) workbook safely and non-destructively."""

    def __init__(self, file_path: Optional[Path | str] = None, auto_backup: bool = True):
        self.file_path = settings.resolve_path(file_path or settings.EXCEL_TRACKER_PATH)
        self.auto_backup = auto_backup
        self._ensure_workbook_exists()

    def _ensure_workbook_exists(self) -> None:
        """Create a styled starter workbook if the user hasn't supplied one yet."""
        if not self.file_path.exists():
            self.file_path.parent.mkdir(parents=True, exist_ok=True)
            logger.info(f"Creating new starter tracking workbook at: {self.file_path}")

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Applications"

            # Header styling
            header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
            header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            border_side = Side(style="thin", color="D9D9D9")
            border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

            for col_idx, col_def in enumerate(DEFAULT_EXCEL_COLUMNS, start=1):
                cell = ws.cell(row=1, column=col_idx, value=col_def.canonical_header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border
                ws.column_dimensions[get_column_letter(col_idx)].width = col_def.width

            ws.row_dimensions[1].height = 28
            ws.freeze_panes = "A2"
            wb.save(self.file_path)

    def _create_backup(self) -> Optional[Path]:
        """Create a timestamped backup before modifying the workbook."""
        if not self.auto_backup or not self.file_path.exists():
            return None

        backup_dir = self.file_path.parent / "backups"
        backup_dir.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = backup_dir / f"{self.file_path.stem}_{timestamp}.xlsx"
        try:
            shutil.copy2(self.file_path, backup_path)
            logger.debug(f"Created workbook backup at: {backup_path}")
            return backup_path
        except Exception as e:
            logger.warning(f"Failed to create backup: {e}")
            return None

    def inspect_workbook(self) -> Dict[str, Any]:
        """Inspect and return metadata about sheets, headers, and rows."""
        wb = openpyxl.load_workbook(self.file_path, data_only=True)
        sheet_names = wb.sheetnames
        ws = wb.active

        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        headers = [str(h) for h in headers if h is not None]

        return {
            "file_path": str(self.file_path),
            "sheet_names": sheet_names,
            "active_sheet": ws.title,
            "headers": headers,
            "total_rows": max(0, ws.max_row - 1),
        }

    def _get_column_mapping(self, ws: openpyxl.worksheet.worksheet.Worksheet) -> Dict[str, int]:
        """Map canonical field keys (e.g. 'company', 'status') to 1-indexed column numbers in the sheet."""
        alias_map = get_column_alias_map()
        col_map: Dict[str, int] = {}

        for col_idx in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col_idx).value
            if val is not None:
                header_str = str(val).strip().lower()
                if header_str in alias_map:
                    canonical_key = alias_map[header_str]
                    col_map[canonical_key] = col_idx

        return col_map

    def find_application_row(
        self,
        ws: openpyxl.worksheet.worksheet.Worksheet,
        col_map: Dict[str, int],
        app: ApplicationRecord
    ) -> Optional[int]:
        """Check for existing matching row using Application ID, URL, or (Company + Title)."""
        app_id_col = col_map.get("application_id")
        job_url_col = col_map.get("job_url")
        company_col = col_map.get("company")
        title_col = col_map.get("job_title")

        for r in range(2, ws.max_row + 1):
            # Check by ID
            if app_id_col:
                val = ws.cell(row=r, column=app_id_col).value
                if val and str(val).strip().upper() == app.application_id.strip().upper():
                    return r

            # Check by Job URL
            if job_url_col and app.job_url:
                val = ws.cell(row=r, column=job_url_col).value
                if val and str(val).strip().lower() == app.job_url.strip().lower():
                    return r

            # Check by normalized Company + Job Title
            if company_col and title_col:
                row_company = str(ws.cell(row=r, column=company_col).value or "").strip().lower()
                row_title = str(ws.cell(row=r, column=title_col).value or "").strip().lower()
                if (
                    row_company
                    and row_title
                    and row_company == app.company.strip().lower()
                    and row_title == app.job_title.strip().lower()
                ):
                    return r

        return None

    def upsert_application(self, app: ApplicationRecord) -> Tuple[str, int]:
        """Insert a new row or update an existing application row in the Excel sheet.

        Returns ('CREATED' or 'UPDATED', row_number).
        """
        self._create_backup()

        wb = openpyxl.load_workbook(self.file_path)
        ws = wb.active
        col_map = self._get_column_mapping(ws)

        # Find existing row or append new row
        existing_row = self.find_application_row(ws, col_map, app)
        target_row = existing_row if existing_row else (ws.max_row + 1)
        action = "UPDATED" if existing_row else "CREATED"

        # Update last_updated timestamp
        app.last_updated = datetime.now(timezone.utc)

        # Map field values
        field_values = {
            "application_id": app.application_id,
            "company": app.company,
            "job_title": app.job_title,
            "location": app.location or "",
            "job_type": app.job_type or "Full-time",
            "match_score": f"{int(app.match_score)}%" if app.match_score is not None else "",
            "status": app.status.value if isinstance(app.status, ApplicationStatus) else str(app.status),
            "resume_used": app.resume_used or "",
            "source": app.source or "Direct",
            "job_url": app.job_url or "",
            "application_url": app.application_url or "",
            "applied_at": app.applied_at.strftime("%Y-%m-%d %H:%M") if app.applied_at else "",
            "last_updated": app.last_updated.strftime("%Y-%m-%d %H:%M"),
            "follow_up_date": app.follow_up_date or "",
            "notes": (app.notes[:150] + "...") if (app.notes and len(app.notes) > 150) else (app.notes or ""),
        }

        # Write cells using column mapping
        for key, val in field_values.items():
            if key in col_map:
                col_idx = col_map[key]
                cell = ws.cell(row=target_row, column=col_idx)
                cell.value = val
                cell.alignment = Alignment(vertical="center")

        # Save workbook safely
        wb.save(self.file_path)
        logger.info(f"Excel Tracker {action} application '{app.company} - {app.job_title}' at row {target_row}")
        return action, target_row

    def get_all_applications(self) -> List[Dict[str, Any]]:
        """Read and return all application rows from the Excel workbook."""
        wb = openpyxl.load_workbook(self.file_path, data_only=True)
        ws = wb.active
        col_map = self._get_column_mapping(ws)

        applications = []
        for r in range(2, ws.max_row + 1):
            row_data = {}
            for key, col_idx in col_map.items():
                row_data[key] = ws.cell(row=r, column=col_idx).value

            if any(row_data.values()):
                applications.append(row_data)

        return applications


excel_tracker = ExcelTracker()
